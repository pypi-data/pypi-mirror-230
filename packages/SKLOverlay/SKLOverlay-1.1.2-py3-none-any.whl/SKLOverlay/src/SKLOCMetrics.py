from os.path import exists, isdir
from numpy import unique
from numpy.random import rand
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
    from openpyxl.utils import column_index_from_string
from sklearn.metrics import *


def Send_To_Excel(scoreList, nickname, path=''):
    print("Sending")

    # Create Dataframe
    scoreListLabels = ['Name', 'Classifier', 'Total_Predictions', 'Unique_Classifications', 'Correct_Predictions',
                       'Incorrect_Predictions', 'Accuracy', 'Balanced_Accuracy', 'Precision', 'Recall', 'F1_Score',
                       'F_Beta Score', 'Phi_Score', 'Hamming_Loss', 'Mean_Absolute_Error', 'Mean_Squared_Error',
                       'Loss', 'Execution Time']

    resultsDF = pd.DataFrame(scoreList, columns=scoreListLabels)

    # Path check and file uniqueness
    if path == 'Default':
        path = ''

    fileName = path + nickname + ".xlsx"
    cloneNum = 0
    while True:
        if exists(fileName):
            cloneNum += 1
            fileName = path + nickname + str(cloneNum) + ".xlsx"
        else:
            break

    resultsDF.to_excel(fileName)

    # ADJUSTING COLUMNS (RIPPED FROM STACK OVERFLOW)
    # Importing the necessary modules
    resultsWB = load_workbook(fileName)
    for sheet_name in resultsWB.sheetnames:
        for column_cells in resultsWB[sheet_name].columns:
            # Max length inside tuple of strings made from data contained in each cell of a column
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                columnBuffer = 1.23
                resultsWB[sheet_name].column_dimensions[new_column_letter].width = new_column_length * columnBuffer
    resultsWB.save(fileName)
    # ADJUSTING COLUMNS (RIPPED FROM STACK OVERFLOW)

    print(resultsDF)


def Save_CM(y_test, y_guess, nickname, classifier, path=''):
    cm = confusion_matrix(y_test, y_guess, labels=classifier.classes_)
    cmDisplay = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=classifier.classes_)
    cmDisplay.plot()

    fileName = path + nickname + "_CM.png"
    cloneNum = 0
    while True:
        if exists(fileName):
            cloneNum += 1
            fileName = path + nickname + "_" + str(cloneNum) + "_CM.png"
        else:
            break

    plt.savefig(fileName, dpi=300)


def Generate_All_Charts(allResults, nickname, path):
    # Create Dataframe
    scoreListLabels = ['Name', 'Classifier', 'Total_Predictions', 'Unique_Classifications', 'Correct_Predictions',
                       'Incorrect_Predictions', 'Accuracy', 'Balanced_Accuracy', 'Precision', 'Recall', 'F1_Score',
                       'F_Beta Score', 'Phi_Score', 'Hamming_Loss', 'Mean_Absolute_Error', 'Mean_Squared_Error',
                       'Loss', 'Execution Time']

    resultsDF = pd.DataFrame(allResults, columns=scoreListLabels)
    colorList = []
    for i in range(resultsDF.shape[0]):
        colorList.append(rand(3))

    for i in range(4, len(scoreListLabels)):
        resultsDF.plot(
            x='Name',
            y=scoreListLabels[i],
            kind='bar',
            color=colorList,
            legend=None)

        # Filenaming
        if path == 'Default':
            path = ''

        fileName = path + nickname + "_" + scoreListLabels[i] + ".png"
        cloneNum = 0
        while True:
            if exists(fileName):
                cloneNum += 1
                fileName = path + nickname + "_" + scoreListLabels[i] + str(cloneNum) + ".png"
            else:
                break

        # Adjusting y-axis to be easier to see differences
        yRange = resultsDF[scoreListLabels[i]]
        bufferPerc = 0.25
        buffer = (yRange.max() - yRange.min()) * bufferPerc + 0.05
        yMin = yRange.min() - buffer
        if yMin < 0:
            yMin = 0
        yMax = yRange.max() + buffer
        plt.ylim((yMin, yMax))

        # Fine tuning visuals
        plt.title(scoreListLabels[i] + " Comparison")
        plt.xlabel("Classifiers")
        plt.ylabel("Scores/Percentages")
        plt.xticks(rotation=30)
        plt.tight_layout()
        for i, v, in enumerate(yRange.tolist()):
            plt.text(i, v, str(v), ha="center")

        plt.savefig(fileName)


def Results_Return(y_test, y_guess, specNickName, classifier, elapsed):
    # Stats
    tolCount = len(y_guess)
    uniCount = len(unique(y_test))
    accCount = accuracy_score(y_test, y_guess, normalize=False)
    accScore = round(accuracy_score(y_test, y_guess) * 100, 2)
    balAccScore = round(balanced_accuracy_score(y_test, y_guess) * 100, 2)
    phiScore = round(matthews_corrcoef(y_test, y_guess), 2)
    hammLossScore = round(hamming_loss(y_test, y_guess) * 100, 2)
    maeScore = round(mean_absolute_error(y_test, y_guess), 2)
    mseScore = round(mean_squared_error(y_test, y_guess), 2)
    zeroOneLossScore = round(zero_one_loss(y_test, y_guess) * 100, 2)
    zeroOneCount = zero_one_loss(y_test, y_guess, normalize=False)
    if uniCount > 2:
        precScore = round(precision_score(y_test, y_guess, average="macro") * 100, 2)
        recScore = round(recall_score(y_test, y_guess, average="macro") * 100, 2)
        f1Score = round(f1_score(y_test, y_guess, average="macro") * 100, 2)
        fBScore = round(fbeta_score(y_test, y_guess, beta=0.5, average="macro") * 100, 2)
    else:
        precScore = round(precision_score(y_test, y_guess, average="binary") * 100, 2)
        recScore = round(recall_score(y_test, y_guess, average="binary") * 100, 2)
        f1Score = round(f1_score(y_test, y_guess, average="binary") * 100, 2)
        fBScore = round(fbeta_score(y_test, y_guess, beta=0.5, average="binary") * 100, 2)

    scoreList = [specNickName, str(classifier), tolCount, uniCount, accCount, zeroOneCount, accScore, balAccScore,
                 precScore, recScore, f1Score, fBScore, phiScore, hammLossScore, maeScore, mseScore, zeroOneLossScore,
                 elapsed]

    return scoreList


def Results(y_test, y_guess, nickname, classifier, elapsed):
    inReview = True
    resultsSent = False
    CMGenerated = False
    exportLocation = 'Default'

    # Stats
    scoreList = Results_Return(y_test, y_guess, nickname, classifier, elapsed)

    tolCount = scoreList[2]
    uniCount = scoreList[3]
    accCount = scoreList[4]
    zeroOneCount = scoreList[5]
    accScore = scoreList[6]
    balAccScore = scoreList[7]
    precScore = scoreList[8]
    recScore = scoreList[9]
    f1Score = scoreList[10]
    fBScore = scoreList[11]
    phiScore = scoreList[12]
    hammLossScore = scoreList[13]
    maeScore = scoreList[14]
    mseScore = scoreList[15]
    zeroOneLossScore = scoreList[16]

    while inReview:
        print(f'| Selected Folder: {exportLocation}')
        print('|                        Core Statistics:                         |')
        print('|                              -----                              |')
        print(f'| Execution Time: {elapsed} seconds')
        print(f'| Total Predictions: {tolCount}')
        print(f'| Correct Predictions: {accCount}')
        print(f'| Incorrect Predictions: {zeroOneCount}')
        print(f'| Accuracy Score: {accScore}%')
        print(f'| Balanced Accuracy Score: {balAccScore}%')
        print(f'| Precision Score: {precScore}%')
        print(f'| Recall Score: {recScore}%')
        print(f'| F1 Score: {f1Score}%')
        print(f'| F-Beta Score: {fBScore}%')  # Precision is twice as important as recall
        print(f'| Matthews (Phi) Coef Score: {phiScore}')
        print(f'| Hamming Loss: {hammLossScore}%')
        print(f'| MAE Loss Score: {maeScore}')
        print(f'| MSE Loss Score: {mseScore}')
        print(f'| Zero-one Loss Score: {zeroOneLossScore}%')
        print('-------------------------------------------------------------------')

        print('|                      Select more options:                       |')
        print('|                              -----                              |')
        print('|  0. No Further Actions Needed                                   |')
        print('|  1. Select Custom Export Location                               |')
        print('|  2. Send Results to Excel Spreadsheet                           |')
        print('|  3. Generate a Confusion Matrix                                 |')
        try:
            choice = int(input('|           Enter here or 0 if you changed your mind: '))
            if choice == 0:
                inReview = False
            elif choice == 1:
                print('|     Type or copy in your absolute filepath here for results.    |')
                print('|    Make sure the folder existed before running this program.    |')
                print('|     Include the slash at the end to specify it is a folder.     |')
                folder = input('| Enter here: ')

                # Being kind to the user
                if folder[-1] != "\\":
                    folder = folder + "\\"

                # Booleans
                endSlash = (folder[-1] == "\\")
                folderExists = isdir(folder)

                if endSlash and folderExists:
                    exportLocation = folder
                else:
                    print('|           Unrecognized option used. Please try again.           |')
                    exportLocation = 'Default'

                print('-------------------------------------------------------------------')
            elif choice == 2:
                if not resultsSent:
                    resultsSent = True
                    scoreList = [scoreList]

                    if exportLocation == 'Default':
                        Send_To_Excel(scoreList, nickname)
                    else:
                        Send_To_Excel(scoreList, nickname, path=exportLocation)
                else:
                    print('|                  Results were already sent :)                   |')
                    print('-------------------------------------------------------------------')
            elif choice == 3:
                if not CMGenerated:
                    CMGenerated = True
                    confirm = 'n'

                    if uniCount > 2:
                        confirm = input('| Are you sure you wish to create a Confusion Matrix (Y/n): ')

                    if confirm != 'Y' or uniCount < 2:
                        if exportLocation == 'Default':
                            Save_CM(y_test, y_guess, nickname, classifier)
                        else:
                            Save_CM(y_test, y_guess, nickname, classifier, path=exportLocation)
                else:
                    print('|                  Matrix was already created :)                  |')
                    print('-------------------------------------------------------------------')
            else:
                print('|          Unrecognized option chosen. Please try again.          |')
                print('-------------------------------------------------------------------')
        except ValueError:
            print('-------------------------------------------------------------------')
            print('|                          ERROR CAUGHT                           |')
            print('| Tip: Unless specifically asked, you will only have to use       |')
            print('|      numbers to move around and make selections. Until further  |')
            print('|      improvements, this error will result in loss of progress   |')
            print('|      or even shutdown of the program. I apologize for any       |')
            print('|      inconvenience.                                             |')
            print('-------------------------------------------------------------------')
