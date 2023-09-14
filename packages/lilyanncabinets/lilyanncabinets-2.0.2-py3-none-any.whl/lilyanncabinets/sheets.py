import openpyxl
import datetime
import os

def createNewSheet():
    today = datetime.date.today()
    date = today.strftime('%d-%m-%Y')

    workbook = openpyxl.Workbook()
    
    directory = 'sheet/'
    if not os.path.exists(directory):
        os.mkdir(directory)


    workbook.save(directory + date +'.xlsx')

def writeToSheet(r, c, count):
    today = datetime.date.today()
    date = today.strftime('%d-%m-%Y')
    directory = 'sheet/'

    if not os.path.exists(directory):
        os.mkdir(directory)

    file_path = os.path.join(directory, date + '.xlsx')

    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    cell = worksheet.cell(row=r, column=c)
    cell.value = count

    workbook.save(file_path)