"""
Author: Joseph Caligiuri
Title: Scan
Version: 2.0.2
Date:9/8/23
"""

import openpyxl
import os
import re



def search_excel_file(sheet_name, keyword, row_offset, column_offset, search_in_column=None, search_in_row=None):
    try:

        if not os.path.exists("data/"):
            os.mkdir("data/")

        path = os.path.join("data/", "db.xlsx")

        if not os.path.exists(path):
            workbook = openpyxl.Workbook()
            workbook.save(path)

        # Load the Excel workbook
        workbook = openpyxl.load_workbook(path)
        
        # Get the desired worksheet by name
        sheet = workbook[sheet_name]
        
        found_cells = []

        data = []

        if search_in_column:
            # Search in a specific column
            column = sheet[search_in_column]
            for cell in column:
                cell_value = str(cell.value)
                if all(keyword in cell_value for keyword in keyword):
                    found_cells.append(cell.coordinate)
                    data.append(cell.offset(row=row_offset, column=column_offset).value)
        elif search_in_row:
            # Search in a specific row
            row = sheet[search_in_row]
            for cell in row:
                cell_value = str(cell.value)
                if all(keyword in cell_value for keyword in keyword):
                    found_cells.append(cell.coordinate)
                    data.append(cell.offset(row=row_offset, column=column_offset).value)
        else:
            # Search in the entire sheet
            for row in sheet.iter_rows():
                for cell in row:
                    cell_value = str(cell.value)
                    if all(keyword in cell_value for keyword in keyword):
                        found_cells.append(cell.coordinate)
                        data.append(cell.offset(row=row_offset, column=column_offset).value)

        if found_cells:
            print(f"Found '{keyword}' in the following cell(s):")
            for cell_coord in found_cells:
                print(cell_coord + f", this is the data: '{data}")
        else:
            print(f"'{keyword}' not found in the specified location(s).")

    except FileNotFoundError:
        print("File not found.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    
    return data


