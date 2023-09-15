#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, generators, nested_scopes, print_function, unicode_literals, with_statement

import json
from functools import update_wrapper

from zenutils import importutils
import openpyxl
import xlsxhelper

from django.core.exceptions import FieldDoesNotExist
from django.conf import settings
from django.urls import reverse
from django.http import HttpResponse
from django.http import JsonResponse
from django.http import HttpResponseRedirect
from django.template.loader import render_to_string
try:
    from urllib.parse import quote as urlquote
except ImportError:
    from django.utils.http import urlquote
from django.utils.translation import gettext_lazy as _
from django.db.models import Q
from django.contrib import messages
from django.contrib import admin
from django.contrib.admin.options import csrf_protect_m
from django.contrib.messages.api import get_messages

from django_middleware_global_request.middleware import get_request

from .models import get_simple_export_header
from .models import get_simple_export_data
from .models import get_simple_export_default_field_settings

def get_power_admin_class():
    """We suggest NOT use PowerAdmin directly.

    from django_power_admin import get_power_admin_class
    PowerAdmin = get_power_admin_class()

    class MyAdmin(PowerAdmin):
        pass
    """
    admin_class = getattr(settings, "DJANGO_POWER_ADMIN_EXTENDED_CLASS", None)
    if admin_class is None:
        admin_class = getattr(settings, "GLOBAL_BASE_MODEL_ADMIN", "django_power_admin.admin.PowerAdmin")
    return importutils.import_from_string(admin_class)

# ---------------------------------------------------------------------------------
# Add extra js & css to admin site
# Work with templates/admin/base.html 
# ---------------------------------------------------------------------------------
EXTRA_JS_KEY = "extra_js"
EXTRA_CSS_KEY = "extra_css"

def add_extra_js(js):
    request = get_request()
    if not hasattr(request, EXTRA_JS_KEY):
        setattr(request, EXTRA_JS_KEY, [])
    getattr(request, EXTRA_JS_KEY).append(js)

def add_extra_css(css):
    request = get_request()
    if not hasattr(request, EXTRA_CSS_KEY):
        setattr(request, EXTRA_CSS_KEY, [])
    getattr(request, EXTRA_CSS_KEY).append(css)


class PowerAdmin(admin.ModelAdmin):
    """
    Powered admin base.

    When you are using PowerAdmin, we suggest that you use `get_power_admin_class` to get the base admin class, 
    so that you can extended all your classes with one setting.
    """

    # ---------------------------------------------------------------------------------
    # required by [Add extra context to admin site]
    # ---------------------------------------------------------------------------------
    add_extra_context_admin_view_name = True
    extra_context_admin_view_name_field = "extra_context_admin_view_name"

    # ---------------------------------------------------------------------------------
    # required by [Add row Toolbar to Changelist]
    # ---------------------------------------------------------------------------------
    add_row_odd_even_class = True
    add_row_index_class = True
    add_row_first_class = True
    add_row_last_class = True

    # ---------------------------------------------------------------------------------
    # required by [Add highlight row to ChangeList]
    # ---------------------------------------------------------------------------------
    CHANGE_LIST_HIGHLIGHT_ROW_IDS_KEY = "_change_list_highlight_row_ids"
    highlight_hover_row = False
    highlight_clicked_row = True


    # required by [Add Sort Buttons To Row Toolbar of Changelist]
    enable_changelist_object_toolbar = True
    show_sortable_buttons = True
    django_sort_field = "display_order"
    django_sort_delta = 100
    django_sort_start = 10000

    # ---------------------------------------------------------------------------------
    # required by [Add read & edit switch button]
    # ---------------------------------------------------------------------------------
    EDIT_FLAG = "_edit_flag"
    EDIT_FLAG_READONLY_VALUE = "0"
    REAL_HAS_CHANGE_PERMISSION_KEY = "_real_has_change_permission"
    REAL_HAS_DELETE_PERMISSION_KEY = "_real_has_delete_permission"

    # required by [add_child_button]
    parent_field_name = "parent"

    # ---------------------------------------------------------------------------------
    # required by [Changelist Toolbar]
    # ---------------------------------------------------------------------------------
    # changelist_toolbar_buttons = []

    # ---------------------------------------------------------------------------------
    # required by [simple share model admin]
    # ---------------------------------------------------------------------------------
    enable_simple_share_model_admin = True
    SIMPLE_EXPORT_MAX_ROWS = 2**20

    # ---------------------------------------------------------------------------------
    # required by [simple export actions]
    # ---------------------------------------------------------------------------------
    # enable_simple_export_action = True
    # enable_simple_export_all_action = True
    # enable_simple_export_filtered_action = True

    
    # ---------------------------------------------------------------------------------
    # required by [extra views]
    # ---------------------------------------------------------------------------------
    # extra_views = ["xx_view_name", "zz_view_name"]

    # #################################################################################
    # Get site settings
    # #################################################################################

    def django_power_admin_get_extra_view_names(self):
        extra_view_names =  list(getattr(self, "extra_views", []))
        extra_view_names += self.django_power_admin_auto_discover_extra_view_names()
        return extra_view_names

    def django_power_admin_get_extra_changelist_parameters(self, request):
        return getattr(self, "extra_changelist_parameters", getattr(self, "extra_change_list_parameters", []))

    # #################################################################################
    # Util methods
    # #################################################################################

    def get_admin_view_url(self, view_name, obj=None):
        """Return str of reversed url

        `view_name` can be in ['changelist', 'change', 'delete', 'history'] or `extra view name` or `extra view func`
        `obj` can be `object_id` str or `obj instance`.
        """
        args = []
        if isinstance(obj, (str, int)):
            args = [obj]
        elif obj:
            args = [obj.pk]
        view_name_full = self.django_power_admin_get_extra_view_name_full(view_name)
        view_name_full_with_namespace = "admin:{view_name_full}".format(
            view_name_full=view_name_full,
        )
        view_url = reverse(view_name_full_with_namespace, args=args)
        return view_url

    def is_change_view(self, request, obj=None):
        """Returns True if it's read/add/change view.
        """
        # alreay set _is_read_view/_is_add_view/_is_change_view flag, use it
        is_read_view_flag = getattr(self, "_is_read_view", False)
        is_add_view_flag = getattr(self, "_is_add_view", False)
        is_change_view_flag = getattr(self, "_is_change_view", False)
        if is_read_view_flag or is_add_view_flag or is_change_view_flag:
            return True
        # try to test the request.path
        if obj:
            change_view_url = self.get_admin_view_url("change", obj)
            if change_view_url == request.path:
                return True
            else:
                return False
        # So we not sure
        return False

    def get_messages(self):
        """Returns (loaded_messages, queued_messages)
        """
        request = get_request()
        mss = get_messages(request)
        _loaded_messages = getattr(mss, "_loaded_messages", [])
        _queued_messages = getattr(mss, "_queued_messages", [])
        return _loaded_messages, _queued_messages

    # #################################################################################
    # Functional methods
    # #################################################################################

    # ---------------------------------------------------------------------------------
    # Set view flags
    # ---------------------------------------------------------------------------------
    def set_view_flags(self, is_changelist_view=False, is_read_view=False, is_add_view=False, is_change_view=False, is_delete_view=False, is_history_view=False, is_move_up_view=False, is_move_down_view=False):
        request = get_request()
    
        setattr(request, "_is_changelist_view", is_changelist_view)
        setattr(request, "_is_read_view", is_read_view)
        setattr(request, "_is_add_view", is_add_view)
        setattr(request, "_is_change_view", is_change_view)
        setattr(request, "_is_delete_view", is_delete_view)
        setattr(request, "_is_history_view", is_history_view)
        setattr(request, "_is_move_up_view", is_move_up_view)
        setattr(request, "_is_move_down_view", is_move_down_view)

        if is_changelist_view:
            setattr(request, "_view_name", "changelist_view")
        if is_read_view:
            setattr(request, "_view_name", "read_view")
        if is_add_view:
            setattr(request, "_view_name", "add_view")
        if is_change_view:
            setattr(request, "_view_name", "change_view")
        if is_delete_view:
            setattr(request, "_view_name", "delete_view")
        if is_history_view:
            setattr(request, "_view_name", "history_view")
        if is_move_up_view:
            setattr(request, "_view_name", "move_up_view")
        if is_move_down_view:
            setattr(request, "_view_name", "move_down_view")
    # ---------------------------------------------------------------------------------
    # Views hooks, There are 6 views
    # changelist view
    # read view
    # add view
    # change view
    # delete view
    # history view
    # ---------------------------------------------------------------------------------
    def pre_all_view(self, **kwargs):
        pass

    def post_all_view(self, **kwargs):
        pass

    def pre_changelist_view(self, request, extra_context=None):
        self.pre_all_view(request=request, extra_context=extra_context)

    def post_changelist_view(self, changelist_view_result, request, extra_context=None):

        # required by: [Add Row Classes to ChangeList]
        extra_js = self.get_changelist_object_row_classes_javascript()
        add_extra_js(extra_js)

        self.post_all_view(view_result=changelist_view_result, changelist_view_result=changelist_view_result, request=request, extra_context=extra_context)
        return changelist_view_result

    def pre_read_view(self, request, object_id=None, form_url='', extra_context=None):
        self.pre_all_view(request=request, object_id=object_id, form_url=form_url, extra_context=extra_context)

    def post_read_view(self, read_view_result, request, object_id=None, form_url='', extra_context=None):
        self.post_all_view(view_result=read_view_result, read_view_result=read_view_result, request=request, object_id=object_id, form_url=form_url, extra_context=extra_context)
        return read_view_result

    def pre_add_view(self, request, form_url='', extra_context=None):
        self.pre_all_view(request=request, form_url=form_url, extra_context=extra_context)

    def post_add_view(self, add_view_result, request, form_url='', extra_context=None):
        self.post_all_view(view_result=add_view_result, add_view_result=add_view_result, request=request, form_url=form_url, extra_context=extra_context)
        return add_view_result

    def pre_change_view(self, request, object_id, form_url='', extra_context=None):
        self.pre_all_view(request=request, object_id=object_id, form_url=form_url, extra_context=extra_context)

    def post_change_view(self, change_view_result, request, object_id, form_url='', extra_context=None):
        self.post_all_view(view_result=change_view_result, change_view_result=change_view_result, request=request, object_id=object_id, form_url=form_url, extra_context=extra_context)
        return change_view_result

    def pre_delete_view(self, request, object_id, extra_context=None):
        self.pre_all_view(request=request, object_id=object_id, extra_context=extra_context)

    def post_delete_view(self, delete_view_result, request, object_id, extra_context=None):
        self.post_all_view(view_result=delete_view_result, delete_view_result=delete_view_result, request=request, object_id=object_id, extra_context=extra_context)
        return delete_view_result

    def pre_history_view(self, request, object_id, extra_context=None):
        self.pre_all_view(request=request, object_id=object_id, extra_context=extra_context)
    
    def post_history_view(self, history_view_result, request, object_id, extra_context=None):
        self.post_all_view(view_result=history_view_result, history_view_result=history_view_result, request=request, object_id=object_id, extra_context=extra_context)
        return history_view_result

    def pre_move_up_view(self, request, object_id, extra_context=None):
        self.pre_all_view(request=request, object_id=object_id, extra_context=extra_context)

    def post_move_up_view(self, move_up_view_result, request, object_id, extra_context=None):
        self.post_all_view(view_result=move_up_view_result, history_view_result=move_up_view_result, request=request, object_id=object_id, extra_context=extra_context)
        return move_up_view_result

    def pre_move_down_view(self, request, object_id, extra_context=None):
        self.pre_all_view(request=request, object_id=object_id, extra_context=extra_context)

    def post_move_down_view(self, move_down_view_result, request, object_id, extra_context=None):
        self.post_all_view(view_result=move_down_view_result, history_view_result=move_down_view_result, request=request, object_id=object_id, extra_context=extra_context)
        return move_down_view_result

    # ---------------------------------------------------------------------------------
    # Add extra context to admin site
    # ---------------------------------------------------------------------------------
    def django_power_admin_get_extra_context(self, request, **kwargs):
        extra_context = {}
        if self.add_extra_context_admin_view_name:
            extra_context[self.extra_context_admin_view_name_field] = getattr(request, "_view_name")
        return extra_context

    def get_changelist_view_extra_context(self, request, extra_context):
        self.django_power_admin_get_changelist_toolbar_buttons(request, extra_context)
        return self.django_power_admin_get_extra_context(request, extra_context=extra_context)
    
    def get_read_view_extra_context(self, request, object_id, form_url, extra_context):
        return self.django_power_admin_get_extra_context(request, object_id=object_id, form_url=form_url, extra_context=extra_context)
    
    def get_add_view_extra_context(self, request, form_url, extra_context):
        return self.django_power_admin_get_extra_context(request, form_url=form_url, extra_context=extra_context)

    def get_change_view_extra_context(self, request, object_id, form_url, extra_context):
        return self.django_power_admin_get_extra_context(request, object_id=object_id, form_url=form_url, extra_context=extra_context)

    def get_delete_view_extra_context(self, request, object_id, extra_context):
        return self.django_power_admin_get_extra_context(request, object_id=object_id, extra_context=extra_context)

    def get_history_view_extra_context(self, request, object_id, extra_context):
        return self.django_power_admin_get_extra_context(request, object_id=object_id, extra_context=extra_context)

    def get_move_up_view_extra_context(self, request, object_id, extra_context):
        return self.django_power_admin_get_extra_context(request, object_id=object_id, extra_context=extra_context)

    def get_move_down_view_extra_context(self, request, object_id, extra_context):
        return self.django_power_admin_get_extra_context(request, object_id=object_id, extra_context=extra_context)

    # ---------------------------------------------------------------------------------
    # Add extra view
    # ---------------------------------------------------------------------------------
    def django_power_admin_get_extra_view_urls(self):
        urlpatterns = []
        for key in self.django_power_admin_get_extra_view_names():
            view = getattr(self, key, None)
            if view and callable(view):
                view_name = getattr(view, "name", key)
                is_object_view = getattr(view, "is_object_view", False)
                url = getattr(view, "url", None)
                if not url:
                    if is_object_view:
                        url = "<path:object_id>/{view_name}/".format(view_name=view_name)
                    else:
                        url = view_name + "/"
                urlpatterns.append(self.django_power_admin_make_urlpattern({
                    "name": self.django_power_admin_get_extra_view_name_full(view_name),
                    "view": self.django_power_admin_extra_view_wrap(view),
                    "url": url,
                }))
        return urlpatterns

    def django_power_admin_make_xlsx_reponse(self, workbook, filename=None):
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        default_filename = "{app_label}.{model_name}.xlsx".format(
            app_label=app_label,
            model_name=model_name,
        )
        filename = filename or getattr(workbook, "_filename", None) or default_filename
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = """attachment; filename="{0}" """.format(urlquote(filename)).strip()
        workbook.save(response)
        return response

    def django_power_admin_extra_view_response_handler(self, view):
        def wrapper(*args, **kwargs):
            result = view(*args, **kwargs)
            if isinstance(result, HttpResponse):
                return result
            elif isinstance(result, openpyxl.Workbook):
                return self.django_power_admin_make_xlsx_reponse(result)
            elif result is None:
                request = get_request()
                return HttpResponseRedirect(request.META.get("HTTP_REFERER"))
            elif isinstance(result, str):
                return HttpResponseRedirect(result)
            else:
                return JsonResponse(result)
        return wrapper

    def django_power_admin_extra_view_wrap(self, view):
        def wrapper(*args, **kwargs):
            result = self.admin_site.admin_view(self.django_power_admin_extra_view_response_handler(view))(*args, **kwargs)
            return result
        wrapper.model_admin = self
        return update_wrapper(wrapper, view)

    def django_power_admin_get_extra_view_name(self, view_name):
        """Returns view's simple name, e.g. changelist, change, delete...
        """
        if not isinstance(view_name, str):
            view_name = getattr(view_name, "name", view_name.__name__)
        return view_name

    def django_power_admin_get_extra_view_name_full(self, view_name):
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        view_name = self.django_power_admin_get_extra_view_name(view_name)
        view_name_full = "{app_label}_{model_name}_{view_name}".format(app_label=app_label, model_name=model_name, view_name=view_name)
        return view_name_full

    def django_power_admin_make_urlpattern(self, config):
        from django.urls import path

        name = config.get("name", None)
        view = config["view"]
        url = config["url"]

        if isinstance(view, str):
            view = getattr(self, view, None)
            if view is None:
                view = importutils.import_from_string(view)
        
        extra_params = {}
        if name:
            extra_params["name"] = name
        return path(url, self.django_power_admin_extra_view_wrap(view), **extra_params)

    def django_power_admin_auto_discover_extra_view_names(self):
        if hasattr(self, "_django_power_admin_auto_discover_extra_view_names"):
            return getattr(self, "_django_power_admin_auto_discover_extra_view_names")
        names = []
        for name in dir(self):
            if (not name.startswith("_")) and (not name in _EXTRA_VIEW_AUTO_DISCOVER_INGORE_PROPS):
                view = getattr(self, name, None)
                if not view is None:
                    if self.is_extra_view(view):
                        names.append(name)
        setattr(self, "_django_power_admin_auto_discover_extra_view_names", names)
        return names

    def is_extra_view(self, view):
        return getattr(view, "is_extra_view", getattr(view, "extra_view", getattr(view, "is_object_view", False))) and callable(view)

    # ---------------------------------------------------------------------------------
    # Add row Toolbar to Changelist
    # ---------------------------------------------------------------------------------
    list_display_links = None
    changelist_object_toolbar_buttons = ["read_button", "change_button", "delete_button"]

    class ChangelistObjectButton(object):
        def __init__(self, obj, modeladmin, button_func_name, button_func=None):
            self.obj = obj
            self.modeladmin = modeladmin
            self.button_func_name = button_func_name
            self.button_func = button_func or getattr(self.modeladmin, self.button_func_name)
            self.prepared = False
        
        def prepare(self):
            if not self.prepared:
                self.prepared = True
                self.label = getattr(self.button_func, "short_description", "")
                self.icon = getattr(self.button_func, "icon", "")
                self.target = getattr(self.button_func, "target", "_self")
                self.classes = getattr(self.button_func, "classes", "")
                if isinstance(self.classes, (list, tuple, set)):
                    self.classes = " ".join(list(self.classes))
                is_object_view = getattr(self.button_func, "is_object_view", None)
                if is_object_view:
                    self.href = self.modeladmin.get_admin_view_url(self.button_func, self.obj)
                else:
                    self.href = self.button_func(self.obj)
                self.add_referer_parameters = getattr(self.button_func, "add_referer_parameters", False)
                if self.add_referer_parameters:
                    if not "?" in self.href:
                        self.href += "?"
                    if not self.href.endswith("?"):
                        self.href += "&"
                    self.href += get_request().GET.urlencode()

        def render(self):
            self.prepare()
            return render_to_string("django_power_admin/admin/ChangelistObjectToolbarButton.html", {
                "request": get_request(),
                "modeladmin": self.modeladmin,
                "obj": self.obj,
                "label": self.label,
                "icon": self.icon,
                "classes": self.classes,
                "target": self.target,
                "href": self.href,
            })

    def get_changelist_object_toolbar_button_names(self, toolbar_name):
        return getattr(self, toolbar_name, [])

    def changelist_object_toolbar(self, obj):
        return self.django_power_admin_make_changelist_object_toolbar(obj, "changelist_object_toolbar_buttons")
    changelist_object_toolbar.short_description = _("Operations")

    def django_power_admin_make_changelist_object_toolbar(self, obj, toolbar_name):
        buttons = []
        button_names = self.get_changelist_object_toolbar_button_names(toolbar_name)
        for button_name in button_names:
            button = self.ChangelistObjectButton(obj, self, button_name)
            buttons.append(button)
        request = get_request()
        return render_to_string("django_power_admin/admin/ChangelistObjectToolbar.html", {
            "buttons": buttons,
            "request": request,
        })

    def read_button(self, obj):
        app_label = obj._meta.app_label
        model_name = obj._meta.model_name
        view_name = "admin:{app_label}_{model_name}_change".format(
            app_label=app_label,
            model_name=model_name,
        )
        url = reverse(view_name, args=[obj.pk]) + "?_edit_flag=0"
        return url
    read_button.short_description = _("Read")
    read_button.classes = "viewlink"

    def change_button(self, obj):
        app_label = obj._meta.app_label
        model_name = obj._meta.model_name
        view_name = "admin:{app_label}_{model_name}_change".format(
            app_label=app_label,
            model_name=model_name,
        )
        url = reverse(view_name, args=[obj.pk]) + "?_edit_flag=1"
        return url
    change_button.short_description = _("Change")
    change_button.classes = "changelink"

    def delete_button(self, obj):
        app_label = obj._meta.app_label
        model_name = obj._meta.model_name
        view_name = "admin:{app_label}_{model_name}_delete".format(
            app_label=app_label,
            model_name=model_name,
        )
        url = reverse(view_name, args=[obj.pk])
        return url
    delete_button.short_description = _("Delete")
    delete_button.classes = "deletelink"

    def add_child_button(self, obj):
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        view_name = "admin:{app_label}_{model_name}_add".format(
            app_label=app_label,
            model_name=model_name,
        )
        url = reverse(view_name) + "?{parent_field_name}_id={pk}".format(parent_field_name=self.parent_field_name, pk=obj.pk)
        return url
    add_child_button.short_description = _("Add Child")
    add_child_button.classes = "addlink"

    # We say view something
    # Now we say read something
    # Keep the old name so that old things goes well
    view_action = read_button
    read_action = read_button
    edit_action = change_button
    edit_button = change_button
    change_action = change_button
    delete_action = delete_button
    add_child_action = add_child_button
    
    # ---------------------------------------------------------------------------------
    # Add Row Classes to ChangeList
    # ---------------------------------------------------------------------------------
    def get_changelist_object_row_classes_javascript(self):
        request = get_request()
        _changelist_instance = getattr(request, "_changelist_instance", None)
        objs = []
        if _changelist_instance:
            objs = _changelist_instance.result_list
        javascripts = []
        index = 0
        for obj in objs:
            class_text = " ".join(self.get_object_row_class(obj, index, objs))
            if class_text:
                javascripts.append("""
                $("#result_list .action-select[value={pk}]").parents("tr").addClass("{class_text}");
                """.format(pk=obj.pk, class_text=class_text))
            index += 1
        if javascripts:
            return """
            $(document).ready(function(){{
                {javascript_text}
            }});
            """.format(javascript_text="\n".join(javascripts))
        else:
            return ""

    def get_object_row_class(self, obj, index, objs):
        classes = []

        # add basic classes to changelist rows
        if self.add_row_odd_even_class:
            if index % 2 == 0:
                classes.append("result_list_row_odd")
            else:
                classes.append("result_list_row_even")
        if self.add_row_index_class:
            classes.append("result_list_row_index_{index}".format(index=index))
        if self.add_row_first_class and index == 0:
            classes.append("result_list_first_row")
        if self.add_row_last_class and index == len(objs) - 1:
            classes.append("result_list_last_row")
        
        # By default, highlight_hover_row is disabled
        # so that it will NOT mass up with move_up&move_down highlights
        # You can enable it by set TheAdminSite's `highlight_hover_row` property to True
        if self.highlight_hover_row:
            classes.append("highlight_hover_row_enabled")
    
        # By default, highlight_clicked_row is enabled
        # You can disable it by set TheAdminSite's `highlight_clicked_row` property to False
        if self.highlight_clicked_row:
            classes.append("highlight_clicked_row_enabled")

        # add `change_list_highlight_row` class to changelist rows
        # required by [Add highlight row to ChangeList]
        highlight_row_ids =  self.get_highlight_row_ids()
        if obj.pk in highlight_row_ids:
            classes.append("change_list_highlight_row")
        
        return classes

    # ---------------------------------------------------------------------------------
    # Add highlight row to ChangeList
    # ---------------------------------------------------------------------------------
    def set_change_list_highlight_rows(self, rows):
        request = get_request()
        request.session[self.CHANGE_LIST_HIGHLIGHT_ROW_IDS_KEY] = json.dumps(rows)

    def get_highlight_row_ids(self):
        request = get_request()
        if hasattr(request, self.CHANGE_LIST_HIGHLIGHT_ROW_IDS_KEY):
            return getattr(request, self.CHANGE_LIST_HIGHLIGHT_ROW_IDS_KEY)
        highlight_row_ids = json.loads(request.session.get(self.CHANGE_LIST_HIGHLIGHT_ROW_IDS_KEY, "[]"))
        setattr(request, self.CHANGE_LIST_HIGHLIGHT_ROW_IDS_KEY, highlight_row_ids)
        if self.CHANGE_LIST_HIGHLIGHT_ROW_IDS_KEY in request.session:
            del request.session[self.CHANGE_LIST_HIGHLIGHT_ROW_IDS_KEY]
        return highlight_row_ids

    # ---------------------------------------------------------------------------------
    # Add Sort Buttons To Row Toolbar of Changelist
    # ---------------------------------------------------------------------------------
    sortable_admin_sort_arrows = [
        "django_power_admin_move_up",
        "django_power_admin_move_down",
    ]

    def django_power_admin_sort_arrows(self, obj):
        return self.django_power_admin_make_changelist_object_toolbar(obj, "sortable_admin_sort_arrows")
    django_power_admin_sort_arrows.short_description = _("Sortable Admin Sortable Arrows")

    def reset_order(self):
        request = get_request()
        ordering = self.get_ordering(request)
        objs = self.model.objects.order_by(*ordering).all()
        index = self.django_sort_start
        for obj in objs:
            setattr(obj, self.django_sort_field, index)
            index += self.django_sort_delta
        return self.model.objects.bulk_update(objs, [self.django_sort_field])

    def django_power_admin_move_up(self, request, object_id, extra_context=None):
        extra_context = extra_context or {}
        self.set_view_flags(is_move_up_view=True)
        extra_context.update(self.get_move_up_view_extra_context(request=request, object_id=object_id, extra_context=extra_context))
        self.pre_move_up_view(request=request, object_id=object_id, extra_context=extra_context)

        object_id = int(object_id)
        obj = self.get_object(request, object_id)
        values = self.get_queryset(request).values_list("id", self.django_sort_field)
        prev_value = None
        self.set_change_list_highlight_rows([obj.pk])
        for index, value in enumerate(values):
            if obj.pk == value[0]:
                if index == 0:
                    prev_id = None
                    prev_value = None
                else:
                    prev_id = values[index-1][0]
                    prev_value = values[index-1][1]
        if prev_id is None:
            messages.error(request, _("This item is already the first."))
        else:
            obj_value = getattr(obj, self.django_sort_field)
            if (not prev_value is None) and (not obj_value is None) and prev_value != obj_value:
                prev = self.model.objects.get(pk=prev_id)
                setattr(obj, self.django_sort_field, prev_value)
                setattr(prev, self.django_sort_field, obj_value)
                self.model.objects.bulk_update([obj, prev], [self.django_sort_field])
            else:
                self.reset_order()
                obj = self.model.objects.get(pk=object_id)
                prev = self.model.objects.get(pk=prev_id)
                obj_value = getattr(obj, self.django_sort_field)
                prev_value = getattr(prev, self.django_sort_field)
                setattr(obj, self.django_sort_field, prev_value)
                setattr(prev, self.django_sort_field, obj_value)
                self.model.objects.bulk_update([obj, prev], [self.django_sort_field])
            messages.success(request, _("Item move up done!"))
    django_power_admin_move_up.is_object_view = True
    django_power_admin_move_up.name = "django_power_admin_move_up"
    django_power_admin_move_up.short_description = _("Move Up")
    django_power_admin_move_up.classes = "sortable_admin_move_up_arrow"
    django_power_admin_move_up.icon = "fa fa-arrow-up"

    def django_power_admin_move_down(self, request, object_id, extra_context=None):
        extra_context = extra_context or {}
        self.set_view_flags(is_move_down_view=True)
        extra_context.update(self.get_move_down_view_extra_context(request=request, object_id=object_id, extra_context=extra_context))
        self.pre_move_down_view(request=request, object_id=object_id, extra_context=extra_context)

        object_id = int(object_id)
        obj = self.get_object(request, object_id)
        values = self.get_queryset(request).values_list("id", self.django_sort_field)
        next_value = None
        self.set_change_list_highlight_rows([obj.pk])
        for index, value in enumerate(values):
            if obj.pk == value[0]:
                if index == len(values) - 1:
                    next_id = None
                    next_value = None
                else:
                    next_id = values[index+1][0]
                    next_value = values[index+1][1]
        if next_id is None:
            messages.error(request, _("This item is already the last."))
        else:
            obj_value = getattr(obj, self.django_sort_field)
            if (not next_value is None) and (not obj_value is None) and next_value != obj_value:
                next = self.model.objects.get(pk=next_id)
                setattr(obj, self.django_sort_field, next_value)
                setattr(next, self.django_sort_field, obj_value)
                self.model.objects.bulk_update([obj, next], [self.django_sort_field])
            else:
                self.reset_order()
                obj = self.model.objects.get(pk=object_id)
                next = self.model.objects.get(pk=next_id)
                obj_value = getattr(obj, self.django_sort_field)
                next_value = getattr(next, self.django_sort_field)
                setattr(obj, self.django_sort_field, next_value)
                setattr(next, self.django_sort_field, obj_value)
                self.model.objects.bulk_update([obj, next], [self.django_sort_field])
            messages.success(request, _("Item move down done!"))
    django_power_admin_move_down.is_object_view = True
    django_power_admin_move_down.name = "django_power_admin_move_down"
    django_power_admin_move_down.short_description = _("Move Down")
    django_power_admin_move_down.classes = "sortable_admin_move_down_arrow"
    django_power_admin_move_down.icon = "fa fa-arrow-down"

    django_power_admin_move_up_arrow = django_power_admin_move_up
    django_power_admin_move_down_arrow = django_power_admin_move_down

    # ---------------------------------------------------------------------------------
    # Add read & edit switch button
    # ---------------------------------------------------------------------------------
    def has_change_permission_real(self, request, obj=None):
        self.has_change_permission(request, obj)
        return getattr(request, self.REAL_HAS_CHANGE_PERMISSION_KEY, None)

    def has_delete_permission_real(self, request, obj=None):
        self.has_delete_permission(request, obj)
        return getattr(request, self.REAL_HAS_DELETE_PERMISSION_KEY, None)

    def has_change_permission(self, request, obj=None):
        """Disable change permission in read_view.

        `change_view` without `edit_flag` is a read view.
        `change_view` without change permission is a read view.

        `change_view` has both `edit_flag` and `change permission` is a change view.
        """
        result = super().has_change_permission(request, obj=obj)
        setattr(request, self.REAL_HAS_CHANGE_PERMISSION_KEY, result)
        if self.is_change_view(request, obj) and request.GET.get(self.EDIT_FLAG, self.EDIT_FLAG_READONLY_VALUE) == self.EDIT_FLAG_READONLY_VALUE:
            return False
        else:
            return result
    
    def has_delete_permission(self, request, obj=None):
        """Disable delete permission in read_view.
        """
        result = super().has_delete_permission(request, obj=obj)
        setattr(request, self.REAL_HAS_DELETE_PERMISSION_KEY, result)
        if self.is_change_view(request, obj) and request.GET.get(self.EDIT_FLAG, self.EDIT_FLAG_READONLY_VALUE) == self.EDIT_FLAG_READONLY_VALUE:
            return False
        else:
            return result

    # ---------------------------------------------------------------------------------
    # Add changelist toolbar buttons
    # ---------------------------------------------------------------------------------
    class ChangelistToolbarButton(object):
        def __init__(self, modeladmin, button_func_name, button_func=None):
            self.modeladmin = modeladmin
            self.button_func_name = button_func_name
            self.button_func = button_func or getattr(self.modeladmin, self.button_func_name)
            self.prepared = False

        def prepare(self):
            if not self.prepared:
                self.prepared = True
                self.label = getattr(self.button_func, "short_description", "")
                self.icon = getattr(self.button_func, "icon", "")
                self.target = getattr(self.button_func, "target", "_self")
                self.classes = getattr(self.button_func, "classes", "")
                if isinstance(self.classes, (list, tuple, set)):
                    self.classes = " ".join(list(self.classes))
                is_extra_view = getattr(self.button_func, "is_extra_view", getattr(self.button_func, "extra_view", False))
                if is_extra_view:
                    self.href = self.modeladmin.get_admin_view_url(self.button_func)
                else:
                    self.href = self.button_func()
                self.add_referer_parameters = getattr(self.button_func, "add_referer_parameters", False)
                if self.add_referer_parameters:
                    if not "?" in self.href:
                        self.href += "?"
                    if not self.href.endswith("?"):
                        self.href += "&"
                    self.href += get_request().GET.urlencode()

        def render(self):
            request = get_request()
            self.prepare()
            return render_to_string("django_power_admin/admin/ChangelistToolbarButton.html", {
                "request": request,
                "modeladmin": self.modeladmin,
                "label": self.label,
                "icon": self.icon,
                "classes": self.classes,
                "href": self.href,
                "target": self.target,
            })
    
    class ChangelistToolbarButtonGroup(object):
        def __init__(self, modeladmin, button_group_func_name, button_group_func=None):
            self.modeladmin = modeladmin
            self.button_group_func_name = button_group_func_name
            self.button_group_func = button_group_func or getattr(self.modeladmin, self.button_group_func_name)
            self.prepared = False
    
        def prepare(self):
            if not self.prepared:
                self.prepared = True
                self.label = getattr(self.button_group_func, "short_description", "")
                self.icon = getattr(self.button_group_func, "icon", "")
                self.classes = getattr(self.button_group_func, "classes", "")
                if isinstance(self.classes, (list, tuple, set)):
                    self.classes = " ".join(list(self.classes))
                self.buttons = []
                button_names = self.button_group_func()
                for button_name in button_names:
                    button = self.modeladmin.ChangelistToolbarButton(self.modeladmin, button_name)
                    button.prepare()
                    self.buttons.append(button)
        
        def render(self):
            request = get_request()
            self.prepare()
            return render_to_string("django_power_admin/admin/ChangelistToolbarButtonGroup.html", {
                "request": request,
                "modeladmin": self.modeladmin,
                "label": self.label,
                "icon": self.icon,
                "classes": self.classes,
                "buttons": self.buttons,
            })

    def django_power_admin_get_changelist_toolbar_buttons(self, request, extra_context):
        # keep old setting `change_list_toolbar_buttons` available
        changelist_toolbar_buttons = getattr(self, "changelist_toolbar_buttons", getattr(self, "change_list_toolbar_buttons", []))
        # add default toolbar buttons
        if self.enable_simple_export_action and self.has_simple_export_permission():
            changelist_toolbar_buttons.append("simple_export_button")
        # make changelist tooolbar buttons
        power_admin_changelist_toolbar_buttons = self.django_power_admin_make_changelist_toolbar_buttons(changelist_toolbar_buttons)
        # add power_admin_changelist_toolbar_buttons to extra_context
        extra_context["power_admin_changelist_toolbar_buttons"] = power_admin_changelist_toolbar_buttons

    def django_power_admin_make_changelist_toolbar_buttons(self, changelist_toolbar_button_names):
        buttons = []
        for name in changelist_toolbar_button_names:
            button_func = getattr(self, name, None)
            if button_func:
                is_group = getattr(button_func, "is_group", False)
                if is_group:
                    button = self.django_power_admin_make_changelist_toolbar_buttongroup(name)
                else:
                    button = self.django_power_admin_make_changelist_toolbar_button(name)
                buttons.append(button)
        return buttons
    
    def django_power_admin_make_changelist_toolbar_button(self, button_func_name):
        return self.ChangelistToolbarButton(self, button_func_name)

    def django_power_admin_make_changelist_toolbar_buttongroup(self, button_group_func_name):
        return self.ChangelistToolbarButtonGroup(self, button_group_func_name)

    # ---------------------------------------------------------------------------------
    # Add simple export button
    # ---------------------------------------------------------------------------------
    @property
    def enable_simple_export_action(self):
        return self.enable_simple_export_all_action or self.enable_simple_export_filtered_action

    @property
    def enable_simple_export_all_action(self):
        if get_request().user.is_superuser: # only super user can export all items
            return True
        else: # normal user can only export items that he can read
            return False

    @property
    def enable_simple_export_filtered_action(self):
        return True

    def has_simple_export_permission(self):
        return self.has_simple_export_all_permission() or self.has_simple_export_filtered_permission()

    def has_simple_export_all_permission(self):
        request = get_request()
        return self.has_view_permission(request)
    
    def has_simple_export_filtered_permission(self):
        request = get_request()
        return self.has_view_permission(request)

    def simple_export_button(self):
        items = []
        if self.enable_simple_export_all_action and self.has_simple_export_all_permission():
            items.append("django_power_admin_simple_export_all")
        if self.enable_simple_export_filtered_action and self.has_simple_export_filtered_permission():
            items.append("django_power_admin_simple_export_filtered")
        return items
    simple_export_button.short_description = _("Export")
    simple_export_button.icon = "fa fa-download"
    simple_export_button.is_group = True

    def django_power_admin_simple_export_all(self, request):
        result_list = self.model.objects.all()
        return self.django_power_admin_do_simple_export(request, result_list)
    django_power_admin_simple_export_all.is_extra_view = True
    django_power_admin_simple_export_all.name = "django_power_admin_simple_export_all"
    django_power_admin_simple_export_all.short_description = _("Export All")
    django_power_admin_simple_export_all.icon = "fa fa-download"

    def django_power_admin_simple_export_filtered(self, request):
        result_list = self.django_power_admin_get_simple_export_items(request)
        return self.django_power_admin_do_simple_export(request, result_list)
    django_power_admin_simple_export_filtered.is_extra_view = True
    django_power_admin_simple_export_filtered.add_referer_parameters = True
    django_power_admin_simple_export_filtered.name = "django_power_admin_simple_export_filtered"
    django_power_admin_simple_export_filtered.short_description = _("Export Filtered")
    django_power_admin_simple_export_filtered.icon = "fa fa-download"


    def django_power_admin_get_simple_export_settings(self):
        return getattr(self, "simple_export_settings", {
            "workbook_template_file": None,
            "filename": None,
            "sheets": [{
                "name": "Sheet1",
                "show-header": True,
                "header-row": 1,
                "start-row-index": 2,
                "filename": None,
                "exclude-fields": [],
            }]
        })

    def django_power_admin_do_simple_export(self, request, result_list):
        settings = self.django_power_admin_get_simple_export_settings()
        # create workbook
        workbook_template_file = settings.get("workbook_template_file", None)
        workbook = xlsxhelper.get_workbook(workbook_template_file)
        if not workbook_template_file:
            del workbook[workbook.active.title]
        filename = settings.get("filename", None)
        setattr(workbook, "_filename", filename)
        # create sheets
        sheet_settings = settings.get("sheets", [])
        sheet_index = 0
        for sheet_setting in sheet_settings:
            sheet_index += 1
            # create a sheet
            name = sheet_setting.get("name", "Sheet{sheet_index}".format(sheet_index=sheet_index))
            try:
                sheet = workbook[name]
            except KeyError:
                sheet = workbook.create_sheet(name)
            # get field settings
            exclude_fields = sheet_setting.get("exclude-fields", [])
            field_settings = sheet_setting.get("fields", get_simple_export_default_field_settings(self.model, exclude_fields))
            # set header
            show_header = sheet_setting.get("show-header", True)
            if show_header:
                header_row = sheet_setting.get("header-row", 0)
                header_values = get_simple_export_header(field_settings, self.model)
                col = 0
                for header_value in header_values:
                    col += 1
                    sheet.cell(header_row, col, header_value)
            # set data
            
            data = get_simple_export_data(result_list, field_settings)
            row = sheet_setting.get("start-row-index", 2)
            for row_data in data:
                col = 0
                for value in row_data:
                    col += 1
                    sheet.cell(row, col, value)
                row += 1
        return workbook

    def django_power_admin_get_simple_export_items(self, request):
        sortable_by = self.get_sortable_by(request)
        ChangeList = self.get_changelist(request)
        cl = ChangeList(
            request,
            self.model,
            [], # list_display
            [], #list_display_links
            self.get_list_filter(request),
            None, #date_hierarchy
            self.get_search_fields(request),
            self.get_list_select_related(request),
            self.SIMPLE_EXPORT_MAX_ROWS, # list_per_page
            False, # list_max_show_all
            [], # list_editable
            self,
            sortable_by,
            self.search_help_text,
        )
        return cl.result_list
    # ---------------------------------------------------------------------------------
    # simple share model
    # ---------------------------------------------------------------------------------
    def django_power_admin_get_simple_share_model_queryset(self, request, queryset):
        if not request.user.is_superuser:
            queryset = self.model.get_user_related_objects(request.user, queryset=queryset)
        return queryset

    # #################################################################################
    # ModelAdmin overrides
    # #################################################################################

    # changelist view entry point
    @csrf_protect_m
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        self.set_view_flags(is_changelist_view=True)
        extra_context.update(self.get_changelist_view_extra_context(request=request, extra_context=extra_context))
        self.pre_changelist_view(request=request, extra_context=extra_context)
        changelist_view_result = super().changelist_view(request=request, extra_context=extra_context)
        changelist_view_result = self.post_changelist_view(changelist_view_result=changelist_view_result, request=request, extra_context=extra_context)
        return changelist_view_result

    # read, add and change view entry point
    @csrf_protect_m
    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        is_read_view = False
        is_add_view = False
        is_change_view = False

        if object_id is None:
            # add view
            is_add_view = True
            self.set_view_flags(is_add_view=True)
            extra_context.update(self.get_add_view_extra_context(request=request, form_url=form_url, extra_context=extra_context))
            self.pre_add_view(request=request, form_url=form_url, extra_context=extra_context)
        else:
            obj = self.get_object(request, object_id)
            if self.has_change_permission(request, obj):
                # change view
                is_change_view = True
                self.set_view_flags(is_change_view=True)
                extra_context.update(self.get_change_view_extra_context(request=request, object_id=object_id, form_url=form_url, extra_context=extra_context))
                self.pre_change_view(request=request, object_id=object_id, form_url=form_url, extra_context=extra_context)
            else:
                # read view
                is_read_view = True
                self.set_view_flags(is_read_view=True)
                extra_context.update(self.get_read_view_extra_context(request=request, object_id=object_id, form_url=form_url, extra_context=extra_context))
                self.pre_read_view(request=request, object_id=object_id, form_url=form_url, extra_context=extra_context)

        changeform_result = super().changeform_view(request, object_id=object_id, form_url=form_url, extra_context=extra_context)

        if is_add_view:
            return self.post_add_view(add_view_result=changeform_result, request=request, form_url=form_url, extra_context=extra_context)
        if is_change_view:
            return self.post_change_view(change_view_result=changeform_result, request=request, object_id=object_id, form_url=form_url, extra_context=extra_context)
        if is_read_view:
            return self.post_read_view(read_view_result=changeform_result, request=request, object_id=object_id, form_url=form_url, extra_context=extra_context)

        # always not reachable here...
        return changeform_result

    # delete view entry point
    @csrf_protect_m
    def delete_view(self, request, object_id, extra_context=None):
        extra_context = extra_context or {}
        self.set_view_flags(is_delete_view=True)
        extra_context.update(self.get_delete_view_extra_context(request=request, object_id=object_id, extra_context=extra_context))
        self.pre_delete_view(request=request, object_id=object_id, extra_context=extra_context)
        delete_view_result = super().delete_view(request=request, object_id=object_id, extra_context=extra_context)
        delete_view_result = self.post_delete_view(delete_view_result=delete_view_result, request=request, object_id=object_id, extra_context=extra_context)
        return delete_view_result

    # history entry point
    def history_view(self, request, object_id, extra_context=None):
        extra_context = extra_context or {}
        self.set_view_flags(is_history_view=True)
        extra_context.update(self.get_history_view_extra_context(request=request, object_id=object_id, extra_context=extra_context))
        self.pre_history_view(request=request, object_id=object_id, extra_context=extra_context)
        history_view_result = super().history_view(request=request, object_id=object_id, extra_context=extra_context)
        history_view_result = self.post_history_view(history_view_result=history_view_result, request=request, object_id=object_id, extra_context=extra_context)
        return history_view_result

    def get_urls(self):
        urlpatterns = self.django_power_admin_get_extra_view_urls()
        urlpatterns += super().get_urls()
        return urlpatterns

    def get_ordering(self, request):
        ordering = super().get_ordering(request)
        if not ordering:
            ordering = [] + list(self.model._meta.ordering or [])
        if not ordering:
            try:
                self.model._meta.get_field(self.django_sort_field)
                ordering = [self.django_sort_field, "-pk"]
            except FieldDoesNotExist:
                ordering = ["-pk"]
        return ordering

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        # required by [simple share model]
        if getattr(self.model, "_is_simple_share_model", False) :
            queryset = self.django_power_admin_get_simple_share_model_queryset(request, queryset)
        return queryset

    def get_changelist_instance(self, request):
        result = super().get_changelist_instance(request)
        setattr(request, "_changelist_instance", result)
        return result

    def get_changelist(self, request, **kwargs):
        ChangeListBase = super().get_changelist(request, **kwargs)
        modeladmin = self
        class IgnoreExtraParametersChangeList(ChangeListBase):
            def __init__(self, request, *args, **kwargs):
                super().__init__(request, *args, **kwargs)
                self._extra_changelist_parameters = {}
                for name in modeladmin.django_power_admin_get_extra_changelist_parameters(request):
                    self._extra_changelist_parameters[name] = self.params.get(name, None)

            def get_queryset(self, request):
                for name in modeladmin.django_power_admin_get_extra_changelist_parameters(request):
                    if name in self.params:
                        del self.params[name]
                queryset = super().get_queryset(request)
                return queryset
        return IgnoreExtraParametersChangeList

    def get_list_display(self, request):
        """
        if `enable_changelist_object_toolbar==True`, append `changelist_object_toolbar` field to `list_display`
        """
        fields = super().get_list_display(request)
        if self.enable_changelist_object_toolbar:
            if not "changelist_object_toolbar" in fields:
                fields = list(fields) + ["changelist_object_toolbar"]
        return fields

    class Media:
        css = {
            "all": [
                "fontawesome/css/all.min.css",
                "jquery-ui/jquery-ui.min.css",
                "django_power_admin/admin/css/AdminFix.css",
                "django_power_admin/admin/css/AdminExtends.css",
                "django_power_admin/admin/css/ChangelistHighlightRow.css",
                "django_power_admin/admin/css/ChangelistObjectToolbar.css",
                "django_power_admin/admin/css/ChangelistToolbar.css",
                "django_power_admin/admin/css/Sortable.css",
            ]
        }
        js = [
            "admin/js/vendor/jquery/jquery.js",
            "jquery-ui/jquery-ui.min.js",
            "django_power_admin/admin/js/ChangelistHighlightRow.js",
            "django_power_admin/admin/js/ChangelistToolbar.js",
            "admin/js/jquery.init.js",
        ]    
    # #################################################################################
    # End of PowerAdmin
    # #################################################################################

_EXTRA_VIEW_AUTO_DISCOVER_INGORE_PROPS = set(dir(PowerAdmin)) - set([
    "django_power_admin_move_up",
    "django_power_admin_move_down",
    "django_power_admin_simple_export_filtered",
    "django_power_admin_simple_export_all",
])
