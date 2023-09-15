import os
from tika import parser
from openpyxl import load_workbook


def convert_pdf(file_path):
    # Convert PDF to plain text
    raw = parser.from_file(file_path)
    return raw["content"]


def convert_doc(file_path):
    # Convert DOC and DOCX to plain text
    raw = parser.from_file(file_path)
    return raw["content"]


def convert_xlsx(file_path):
    # Convert XLSX to plain text
    wb = load_workbook(file_path)
    text = ""
    for sheet in wb:
        for row in sheet.iter_rows():
            for cell in row:
                text += str(cell.value) + " "
            text += "\n"
    return text


def convert_txt(file_path):
    # Read plain text
    with open(file_path, "r") as f:
        return f.read()


def save_as_txt(content, output_path):
    # Save content as a TXT file
    with open(output_path, "w") as f:
        f.write(content)