import decimal
import os.path
import traceback
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5 import QtCore
from PyQt5.QtCore import *
import numpy as np
import datetime as dt
from datetime import datetime, timedelta
from dateutil.relativedelta import *
import sys
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import time
from mysqlquerys import connect

np.set_printoptions(linewidth=250)
__version__ = 'V5'


def calculate_last_day_of_month(mnth):
    if mnth < 12:
        lastDayOfMonth = datetime(datetime.now().year, mnth + 1, 1) - timedelta(days=1)
        lastDayOfMonth = lastDayOfMonth.day
    elif mnth == 12:
        lastDayOfMonth = 31
    return lastDayOfMonth


class Income:
    def __init__(self, row, tableHead):
        self.row = row
        # print(row)
        self.tableHead = tableHead

    def calculate_income(self):
        base = 0
        plus = []
        minus = []
        for i, col in enumerate(self.tableHead):
            # print(i, col, self.row[i])
            if col == 'value':
                base = self.row[i]
            elif '%' in col and self.row[i] > 0:
                proc = self.row[i] / 100
                plus.append(proc)
            elif '%' in col and self.row[i] < 0:
                proc = self.row[i] / 100
                minus.append(proc)

        plus_val = 0
        for p in plus:
            val = p * base
            plus_val += val

        base_plus = base + plus_val
        # print('base_plus', base_plus)

        minus_val = 0
        for m in minus:
            val = m * base_plus
            minus_val += val
        final_value = base_plus + minus_val
        # print('base', base)
        # print('plus_val', plus_val)
        # print('minus_val', minus_val)
        # print('final_value', final_value)

        return final_value


class Cheltuiala:
    def __init__(self, row, tableHead):
        self.tableHead = tableHead
        self.id = None
        self.name = None
        self.conto = None
        self.value = None
        self.valid_from = None
        self.valid_to = None
        self.freq = None
        self.pay_day = None
        # self.post_pay = None
        self.auto_ext = None
        self.table = None
        self.read_row(row)

    def read_row(self, row):
        # print(self.tableHead)
        # print(row)

        idIndx = self.tableHead.index('id')
        nameIndx = self.tableHead.index('name')
        contoIndx = self.tableHead.index('myconto')
        valueIndx = self.tableHead.index('value')
        validFromIndx = self.tableHead.index('valid_from')
        validToIndx = self.tableHead.index('valid_to')
        freqIndx = self.tableHead.index('freq')
        payDayIndx = self.tableHead.index('pay_day')
        # postPayIndx = self.tableHead.index('post_pay')
        autoExtIndx = self.tableHead.index('auto_ext')
        self.id = row[idIndx]
        self.name = row[nameIndx]
        self.myconto = row[contoIndx]
        self.value = row[valueIndx]
        self.valid_from = row[validFromIndx]
        self.valid_to = row[validToIndx]
        self.freq = row[freqIndx]
        self.pay_day = row[payDayIndx]
        # print('1111', self.pay_day)

        if self.pay_day is None:
            # print('??????')
            self.pay_day = calculate_last_day_of_month(self.valid_from.month)
        # print('payday.....', self.pay_day)
        # sys.exit()
        # post_pay = row[postPayIndx]
        # if post_pay is None or post_pay == 0:
        #     self.post_pay = False
        # else:
        #     self.post_pay = True

        auto_ext = row[autoExtIndx]
        if auto_ext is None or auto_ext == 0:
            self.auto_ext = False
        else:
            self.auto_ext = True

    # @table.setter
    def set_table(self, table_name):
        self.table = table_name

    @property
    def first_payment(self):
        try:
            first_payment = datetime(self.valid_from.year, self.valid_from.month, self.pay_day)
        except:
            # print(self.id, self.table, self.name)
            # print(self.valid_from.year, self.valid_from.month, self.pay_day)
            # first_payment = calculate_last_day_of_month(selectedStartDate.month)
            first_payment = datetime(self.valid_from.year, self.valid_from.month, calculate_last_day_of_month(self.valid_from.month))
        return first_payment

    def list_of_payments_valid_from_till_selected_end_date(self, selectedEndDate):
        list_of_payments_till_selected_end_date = []
        if self.valid_from <= self.first_payment.date() <= selectedEndDate:
            list_of_payments_till_selected_end_date.append(self.first_payment)

        next_payment = self.first_payment + relativedelta(months=self.freq)
        if next_payment.day != self.pay_day:
            try:
                next_payment = datetime(next_payment.year, next_payment.month, self.pay_day)
            except:
                next_payment = datetime(next_payment.year, next_payment.month, calculate_last_day_of_month(next_payment.month))
        if self.valid_from <= next_payment.date() <= selectedEndDate:
            list_of_payments_till_selected_end_date.append(next_payment)

        while next_payment.date() <= selectedEndDate:
            next_payment = next_payment + relativedelta(months=self.freq)
            if next_payment.day != self.pay_day:
                try:
                    next_payment = datetime(next_payment.year, next_payment.month, self.pay_day)
                except:
                    next_payment = datetime(next_payment.year, next_payment.month,
                                            calculate_last_day_of_month(next_payment.month))
            if self.valid_from <= next_payment.date() <= selectedEndDate:
                list_of_payments_till_selected_end_date.append(next_payment)
        return list_of_payments_till_selected_end_date

    def cut_all_before_selectedStartDate(self, lista, selectedStartDate):
        new_list = []
        for date in lista:
            if date.date() >= selectedStartDate:
                new_list.append(date)
        return new_list

    def cut_all_after_valid_to(self, lista):
        new_list = []
        for date in lista:
            if date.date() <= self.valid_to:
                new_list.append(date)
        return new_list

    def calculate_payments_in_interval(self, selectedStartDate, selectedEndDate):
        list_of_payments_valid_from_till_selected_end_date = self.list_of_payments_valid_from_till_selected_end_date(selectedEndDate)
        # print(20*'*')
        # for i in list_of_payments_valid_from_till_selected_end_date:
        #     print(i)
        # print(20*'*')

        list_of_payments_selected_start_date_till_selected_end_date = self.cut_all_before_selectedStartDate(list_of_payments_valid_from_till_selected_end_date, selectedStartDate)
        # print(20*'*')
        # for i in list_of_payments_selected_start_date_till_selected_end_date:
        #     print(i)
        # print(20*'*')

        if self.valid_to and self.valid_to < selectedEndDate and not self.auto_ext:
            list_of_payments_selected_start_date_till_selected_end_date = self.cut_all_after_valid_to(list_of_payments_selected_start_date_till_selected_end_date)
            # print(20*'*')
            # for i in list_of_payments_selected_start_date_till_selected_end_date:
            #     print(i)
            # print(20*'*')

        return list_of_payments_selected_start_date_till_selected_end_date

    @property
    def first_payment_date(self):
        first_payment_date = datetime(self.valid_from.year, self.valid_from.month, self.pay_day)
        return first_payment_date

    @property
    def payments_for_interval(self):
        return self.pfi

    @payments_for_interval.setter
    def payments_for_interval(self, payments_days):
        self.pfi= payments_days


class CheltuieliPlanificate:
    def __init__(self):
        self.ini_file = r"D:\Python\MySQL\database.ini"
        self.data_base_name = 'cheltuieli'
        self.tableHead = ['id', 'name', 'value', 'myconto', 'freq', 'pay_day', 'valid_from', 'valid_to', 'auto_ext']#, 'post_pay'
        self.myAccountsTable = connect.Table(self.ini_file, 'myfolderstructure', 'banca')
        self.myContos = self.myAccountsTable.returnColumn('name')
        try:
            self.dataBase = connect.DataBase(self.ini_file, self.data_base_name)
        except FileNotFoundError as err:
            iniFile, a = QFileDialog.getOpenFileName(None, 'Open data base configuration file', os.getcwd(), "data base config files (*.ini)")
            if os.path.exists(iniFile):
                self.dataBase = connect.DataBase(iniFile, self.data_base_name)
            # ctypes.windll.user32.MessageBoxW(0, "Your text", "Your title", 1)
        except Exception as err:
            print(traceback.format_exc())

    # def get_one_time_transactions(self, all_chelt):
    #     print(self.one_time_trans.columnsNames)
    #     for row in self.one_time_trans.data:
    #         print(20*'ä', row)
    #         for col in self.tableHead:
    #             if col in self.one_time_trans.columnsNames:
    #                 print('BINGO', col)
    #                 print(row[self.one_time_trans.columnsNames.index(col)])
    #             else:
    #                 print(20*'#', col)
    #     return all_chelt

    def get_all_sql_vals(self):
        print(sys._getframe().f_code.co_name)
        all_chelt = []
        for table in self.dataBase.tables:
            self.dataBase.active_table = table
            check = all(item in list(self.dataBase.active_table.columnsProperties.keys()) for item in self.tableHead)
            if check:
                vals = self.dataBase.active_table.returnColumns(self.tableHead)
                for row in vals:
                    row = list(row)
                    chelt = Cheltuiala(row, self.tableHead)
                    chelt.set_table(table)
                    all_chelt.append(chelt)
        return all_chelt

    def filter_dates(self, all_chelt, selectedStartDate, selectedEndDate):
        print(sys._getframe().f_code.co_name, selectedStartDate, selectedEndDate)
        remaining = []
        for chelt in all_chelt:
            # print(chelt.table, chelt.name, chelt.id, chelt.pay_day)
            payments_in_interval = chelt.calculate_payments_in_interval(selectedStartDate, selectedEndDate)
            # print(payments_in_interval)
            # if chelt.name == 'Steuererklärung_2022':
            #     print(chelt.table, chelt.name, chelt.id, chelt.pay_day, payments_in_interval)
            if isinstance(payments_in_interval, list):
                chelt.payments_for_interval = payments_in_interval
                # print(chelt.table, chelt.name, chelt.id, chelt.pay_day, chelt.payments_for_interval)
                if chelt.payments_for_interval:
                    remaining.append(chelt)
        return remaining

    def filter_conto(self, chelt_list, conto):
        remaining = []
        for ch in chelt_list:
            if conto == 'all' and ch.table != 'intercontotrans':
                remaining.append(ch)
            elif ch.myconto == conto:
                remaining.append(ch)

        return remaining

    def split_expenses_income(self, chelt):
        arr_expenses = []
        arr_income = []
        for ch in chelt:
            if ch.value == 0:
                continue
            for payment_day in ch.payments_for_interval:
                if ch.name == 'ERA-Leistungszul':
                    print('****', ch.name, ch.value, ch.payments_for_interval)

                if ch.value and ch.value > 0:
                    incomeTable = connect.Table(self.ini_file, self.data_base_name, ch.table)
                    full_row = list(incomeTable.returnRowsWhere(('id', ch.id))[0])
                    venit_instance = Income(full_row, incomeTable.columnsNames)
                    ch.value = venit_instance.calculate_income()

                variables = vars(ch)
                row = [ch.table]
                for col in self.tableHead:
                    val = variables[col]
                    row.append(val)
                row.append(payment_day)
                if ch.value and ch.value > 0:
                    arr_income.append(row)
                else:
                    arr_expenses.append(row)
        arr_expenses = np.atleast_2d(arr_expenses)
        arr_income = np.atleast_2d(arr_income)
        self.tableHead.insert(0, 'table')
        self.tableHead.append('payDay')
        return arr_expenses, arr_income

    def prepareTablePlan(self, conto, selectedStartDate, selectedEndDate):
        print(sys._getframe().f_code.co_name)

        all_chelt = self.get_all_sql_vals()
        # for i in all_chelt:
        #     print(i.freq)
        # all_chelt = self.get_one_time_transactions(all_chelt)

        chelt_in_time_interval = self.filter_dates(all_chelt, selectedStartDate, selectedEndDate)
        # for chelt in chelt_in_time_interval:
        #     print(chelt.table, chelt.name, chelt.id, chelt.pay_day, chelt.conto, chelt.payments_for_interval)

        chelt_after_contofilter = self.filter_conto(chelt_in_time_interval, conto)
        # for chelt in chelt_after_contofilter:
        #     print(chelt.table, chelt.name, chelt.id, chelt.pay_day, chelt.conto, chelt.payments_for_interval)

        expenses, income = self.split_expenses_income(chelt_after_contofilter)
        if expenses.shape == (1, 0):
            expenses = np.empty((0, len(self.tableHead)))
        if income.shape == (1, 0):
            income = np.empty((0, len(self.tableHead)))

        # for row in income:
        #     print(row)
        return self.tableHead, expenses, income


class MyApp(QMainWindow, CheltuieliPlanificate):
    def __init__(self):
        super(MyApp, self).__init__()
        path2src, pyFileName = os.path.split(__file__)
        uiFileName = 'chelt_plan.ui'
        path2GUI = os.path.join(path2src, 'GUI', uiFileName)
        Ui_MainWindow, QtBaseClass = uic.loadUiType(path2GUI)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        title = '{}_{}'.format(pyFileName, __version__)
        self.setWindowTitle(title)

        self.populateCBConto()
        self.populateCBMonths()
        self.populateDatesInterval()
        self.get_table_info()

        self.ui.cbActiveConto.currentIndexChanged.connect(self.get_table_info)
        self.ui.CBMonths.currentIndexChanged.connect(self.populateDatesInterval)
        self.ui.DEFrom.dateTimeChanged.connect(self.get_table_info)
        self.ui.DEBis.dateTimeChanged.connect(self.get_table_info)
        self.ui.planTable.horizontalHeader().sectionClicked.connect(self.sortPlan)
        self.ui.PB_plotTablePie.clicked.connect(self.plotTablePie)
        self.ui.PB_plotNamePie.clicked.connect(self.plotNamePie)
        self.ui.PB_Plot.clicked.connect(self.plotGraf)
        self.ui.PB_export.clicked.connect(self.export)

    def get_table_info(self):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        displayTableHead = ['table', 'name', 'value', 'myconto', 'payDay', 'freq']
        selectedStartDate = self.ui.DEFrom.date().toPyDate()
        selectedEndDate = self.ui.DEBis.date().toPyDate()
        # tableHead, payments4Interval, income = self.prepareTablePlan(self.ui.cbActiveConto.currentText(), selectedStartDate, selectedEndDate)
        app = CheltuieliPlanificate()
        tableHead, payments4Interval, income = app.prepareTablePlan(self.ui.cbActiveConto.currentText(), selectedStartDate, selectedEndDate)

        self.populateExpensesPlan(tableHead, payments4Interval, displayTableHead)
        self.populateTree(tableHead, payments4Interval)
        self.populateIncomePlan(tableHead, income, displayTableHead)
        self.totals()

    def export(self):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        expName, _ = QFileDialog.getSaveFileName(self, "Save file", "", "Excel Files (*.xlsx)")
        worksheets = [('Complete', datetime(datetime.now().year, 1, 1),datetime(datetime.now().year, 12, 31))]
        for mnth in range(1, 13):
            firstDayOfMonth = datetime(datetime.now().year, mnth, 1)
            if mnth != 12:
                lastDayOfMonth = datetime(datetime.now().year, mnth+1, 1) - timedelta(days=1)
            else:
                lastDayOfMonth = datetime(datetime.now().year + 1, 1, 1) - timedelta(days=1)

            tup = (firstDayOfMonth.strftime("%B"), firstDayOfMonth, lastDayOfMonth)
            worksheets.append(tup)

        wb = Workbook()
        ws = wb.active
        for mnth, firstDayOfMonth, lastDayOfMonth in worksheets:
            # print(mnth, firstDayOfMonth, lastDayOfMonth)
            if mnth == 'Complete':
                ws.title = mnth
            else:
                wb.create_sheet(mnth)
            ws = wb[mnth]
            self.ui.DEFrom.setDate(QDate(firstDayOfMonth))
            self.ui.DEBis.setDate(QDate(lastDayOfMonth))
            self.prepareTablePlan()

            planExpenseTable, planExpenseTableHead = self.readPlanExpenses()
            cheltData = np.insert(planExpenseTable, 0, planExpenseTableHead, 0)

            for i, row in enumerate(cheltData):
                for j, col in enumerate(row):
                    ws.cell(row=i + 1, column=j + 1).value = cheltData[i][j]

            firstRow = 1
            firstCol = get_column_letter(1)
            lastRow = len(cheltData)
            lastCol = get_column_letter(len(cheltData[0]))

            table_title = '{}_{}'.format('chelt', mnth )
            new_text = ('{}{}:{}{}'.format(firstCol, firstRow, lastCol, lastRow))
            tab = Table(displayName=table_title, ref=new_text)
            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)
            ws.cell(row=lastRow + 1, column=1).value = 'Total Number of Expenses'
            ws.cell(row=lastRow + 1, column=2).value = self.ui.LEtotalNoOfTransactions.text()
            ws.cell(row=lastRow + 2, column=1).value = 'Total Expenses'
            ws.cell(row=lastRow + 2, column=2).value = self.ui.LEtotalValue.text()
            #######income

            planIncomeTable, planIncomeTableHead = self.readPlanIncome()
            incomeData = np.insert(planIncomeTable, 0, planIncomeTableHead, 0)
            firstRow = lastRow + 5
            firstCol = get_column_letter(1)
            lastRow = firstRow + len(incomeData)
            lastCol = get_column_letter(len(incomeData[0]))

            for i, row in enumerate(incomeData):
                for j, col in enumerate(row):
                    ws.cell(row=i + firstRow, column=j + 1).value = incomeData[i][j]

            table_title = '{}_{}'.format('income', mnth )
            new_text1 = ('{}{}:{}{}'.format(firstCol, firstRow, lastCol, lastRow))
            tab = Table(displayName=table_title, ref=new_text1)
            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)
            ws.cell(row=lastRow + 1, column=1).value = 'Total Number of Incomes'
            ws.cell(row=lastRow + 1, column=2).value = self.ui.LEtotalNoOfIncome.text()
            ws.cell(row=lastRow + 2, column=1).value = 'Total Income'
            ws.cell(row=lastRow + 2, column=2).value = self.ui.LEtotalIncome.text()

        wb.save(expName)

    def populateCBMonths(self):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        self.ui.CBMonths.addItem('interval')
        months = [dt.date(2000, m, 1).strftime('%B') for m in range(1, 13)]
        for month in months:
            self.ui.CBMonths.addItem(month)

    def populateCBConto(self):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        self.ui.cbActiveConto.addItem('all')
        self.ui.cbActiveConto.addItems(self.myContos)

    def populateDatesInterval(self):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        startDate = QDate(datetime.now().year, datetime.now().month, datetime.now().day)
        if datetime.now().month != 12:
            mnth = datetime.now().month + 1
            lastDayOfMonth = datetime(datetime.now().year, mnth, 1) - timedelta(days=1)
        else:
            lastDayOfMonth = datetime(datetime.now().year + 1, 1, 1) - timedelta(days=1)

        if self.ui.CBMonths.currentText() != 'interval':
            mnth = datetime.strptime(self.ui.CBMonths.currentText(), "%B").month
            # print('****', mnth)
            # if mnth == 1:
            #     startDate = datetime(datetime.now().year - 1, 12, 30)
            # elif mnth == 2:
            #     startDate = datetime(datetime.now().year, mnth-1, 28)
            # else:
            #     startDate = datetime(datetime.now().year, mnth-1, 30)
            #
            # lastDayOfMonth = datetime(datetime.now().year, mnth, 29)

            startDate = datetime(datetime.now().year, mnth, 1)
            if mnth != 12:
                lastDayOfMonth = datetime(datetime.now().year, mnth+1, 1) - timedelta(days=1)
            else:
                lastDayOfMonth = datetime(datetime.now().year + 1, 1, 1) - timedelta(days=1)

            startDate = startDate - timedelta(days=2)
            lastDayOfMonth = lastDayOfMonth - timedelta(days=2)

            startDate = QDate(startDate)
            lastDayOfMonth = QDate(lastDayOfMonth)

        self.ui.DEFrom.setDate(startDate)
        self.ui.DEBis.setDate(lastDayOfMonth)

        self.ui.DEFrom.setCalendarPopup(True)
        self.ui.DEBis.setCalendarPopup(True)

    def populateTree(self, tableHead, table):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        self.ui.TWmnthVSIrreg.clear()
        self.ui.TWmnthVSIrreg.setHeaderLabels(['freq', 'name', 'value'])
        monthly_level = QTreeWidgetItem(self.ui.TWmnthVSIrreg)
        monthly_level.setText(0, 'Monthly')
        irregular_level = QTreeWidgetItem(self.ui.TWmnthVSIrreg)
        irregular_level.setText(0, 'Irregular')
        monthlyIndx = np.where(table[:, tableHead.index('freq')] == 1)
        monthly = table[monthlyIndx]
        for mnth in monthly:
            mnth_item_level = QTreeWidgetItem(monthly_level)
            mnth_item_level.setText(1, mnth[tableHead.index('name')])
            mnth_item_level.setText(2, str(round(mnth[tableHead.index('value')])))

        totalMonthly = table[monthlyIndx,tableHead.index('value')][0]
        monthly_level.setText(1, 'Total')
        monthly_level.setText(2, str(round(sum(totalMonthly), 2)))

        irregIndx = np.where(table[:, tableHead.index('freq')] != 1)
        irregular = table[irregIndx]
        for irr in irregular:
            irr_item_level = QTreeWidgetItem(irregular_level)
            irr_item_level.setText(1, irr[tableHead.index('name')])
            irr_item_level.setText(2, str(round(irr[tableHead.index('value')], 2)))

        totalIrreg = table[irregIndx,tableHead.index('value')][0]
        irregular_level.setText(1, 'Total')
        irregular_level.setText(2, str(round(sum(totalIrreg), 2)))

    def populateExpensesPlan(self, tableHead, table, displayTableHead=None):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        #
        # print(tableHead)
        # print(table)

        if displayTableHead:
            tableHead, table = self.convert_to_display_table(tableHead, table, displayTableHead)

        self.ui.planTable.setColumnCount(len(tableHead))
        self.ui.planTable.setHorizontalHeaderLabels(tableHead)
        self.ui.planTable.setRowCount(table.shape[0])
        for col in range(table.shape[1]):
            for row in range(table.shape[0]):
                if isinstance(table[row, col], int) or isinstance(table[row, col], float):
                    item = QTableWidgetItem()
                    item.setData(QtCore.Qt.DisplayRole, table[row, col])
                elif isinstance(table[row, col], decimal.Decimal):
                    val = float(table[row, col])
                    item = QTableWidgetItem()
                    item.setData(QtCore.Qt.DisplayRole, val)
                else:
                    item = QTableWidgetItem(str(table[row, col]))
                self.ui.planTable.setItem(row, col, item)

        if table.shape[1] > 0:
            self.populate_expenses_summary(tableHead, table)

    def convert_to_display_table(self, tableHead, table, displayTableHead):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        newTableData = np.empty([table.shape[0], len(displayTableHead)], dtype=object)
        for i, col in enumerate(displayTableHead):
            indxCol = tableHead.index(col)
            newTableData[:,i] = table[:, indxCol]

        return displayTableHead, newTableData

    def populate_expenses_summary(self, tableHead, table):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        allValues = table[:, tableHead.index('value')]
        if None in allValues:
            allValues = allValues[allValues != np.array(None)]
        totalVal = round(sum(allValues.astype(float)), 2)
        self.ui.LEtotalNoOfTransactions.setText(str(len(table)))
        self.ui.LEtotalValue.setText(str(totalVal))

        indxMonthly = np.where(table[:,tableHead.index('freq')] == 1)[0]
        monthly = table[indxMonthly, tableHead.index('value')]
        if None in monthly:
            monthly = monthly[monthly != np.array(None)]
        totalMonthly = round(sum(monthly.astype(float)), 2)
        self.ui.LEnoOfMonthly.setText(str(monthly.shape[0]))
        self.ui.LEtotalMonthly.setText(str(totalMonthly))

        indxIrregular = np.where(table[:,tableHead.index('freq')] != 1)[0]
        irregular = table[indxIrregular, tableHead.index('value')]
        if None in irregular:
            irregular = irregular[irregular != np.array(None)]
        totalIrregular = round(sum(irregular.astype(float)), 2)
        self.ui.LEnoOfIrregular.setText(str(irregular.shape[0]))
        self.ui.LEirregular.setText(str(totalIrregular))

    def populateIncomePlan(self, tableHead, table, displayTableHead=None):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        if displayTableHead:
            tableHead, table = self.convert_to_display_table(tableHead, table, displayTableHead)

        self.ui.planTableIncome.setColumnCount(len(tableHead))
        self.ui.planTableIncome.setHorizontalHeaderLabels(tableHead)
        self.ui.planTableIncome.setRowCount(table.shape[0])
        for col in range(table.shape[1]):
            for row in range(table.shape[0]):
                if isinstance(table[row, col], int) or isinstance(table[row, col], float):
                    item = QTableWidgetItem()
                    item.setData(QtCore.Qt.DisplayRole, table[row, col])
                elif isinstance(table[row, col], decimal.Decimal):
                    val = float(table[row, col])
                    item = QTableWidgetItem()
                    item.setData(QtCore.Qt.DisplayRole, val)
                else:
                    item = QTableWidgetItem(str(table[row, col]))
                self.ui.planTableIncome.setItem(row, col, item)

        if table.shape[1] > 0:
            self.populate_income_summary(tableHead, table)

    def populate_income_summary(self, tableHead, table):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        allValues = table[:, tableHead.index('value')]
        if None in allValues:
            allValues = allValues[allValues != np.array(None)]
        # for i in allValues:
        #     print(i, type(i))
        totalVal = sum(allValues.astype(float))
        totalVal = round((totalVal), 2)
        self.ui.LEtotalNoOfIncome.setText(str(len(table)))
        self.ui.LEtotalIncome.setText(str(totalVal))

    def totals(self):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        if self.ui.LEtotalNoOfTransactions.text():
            expensesTrans = int(self.ui.LEtotalNoOfTransactions.text())
        else:
            expensesTrans = 0
        if self.ui.LEtotalNoOfIncome.text():
            incomeTrans = int(self.ui.LEtotalNoOfIncome.text())
        else:
            incomeTrans = 0

        if self.ui.LEtotalValue.text():
            expenses = float(self.ui.LEtotalValue.text())
        else:
            expenses = 0
        if self.ui.LEtotalIncome.text():
            income = float(self.ui.LEtotalIncome.text())
        else:
            income = 0

        trans = expensesTrans + incomeTrans
        total = round(expenses + income, 2)

        self.ui.LEtotalNo.setText(str(trans))
        self.ui.LEtotalVa.setText(str(total))

    def sortPlan(self, logical_index):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        header = self.ui.planTable.horizontalHeader()
        order = Qt.DescendingOrder
        if not header.isSortIndicatorShown():
            header.setSortIndicatorShown(True)
        elif header.sortIndicatorSection() == logical_index:
            order = header.sortIndicatorOrder()
        header.setSortIndicator(logical_index, order)
        self.ui.planTable.sortItems(logical_index, order)

    def readPlanExpenses(self):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        rows = self.ui.planTable.rowCount()
        cols = self.ui.planTable.columnCount()
        planExpenseTable = np.empty((rows, cols), dtype=object)
        planExpenseTableHead = []
        for row in range(rows):
            for column in range(cols):
                cell = self.ui.planTable.item(row, column)
                planExpenseTable[row, column] = cell.text()
                colName = self.ui.planTable.horizontalHeaderItem(column).text()
                if colName not in planExpenseTableHead:
                    planExpenseTableHead.append(colName)

        return planExpenseTable, planExpenseTableHead

    def readPlanIncome(self):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        rows = self.ui.planTableIncome.rowCount()
        cols = self.ui.planTableIncome.columnCount()
        planIncomeTable = np.empty((rows, cols), dtype=object)
        planIncomeTableHead = []
        for row in range(rows):
            for column in range(cols):
                cell = self.ui.planTableIncome.item(row, column)
                planIncomeTable[row, column] = cell.text()
                colName = self.ui.planTableIncome.horizontalHeaderItem(column).text()
                if colName not in planIncomeTableHead:
                    planIncomeTableHead.append(colName)

        return planIncomeTable, planIncomeTableHead

    def plotTablePie(self):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        realExpenseTable, realExpenseTableHead = self.readPlanExpenses()
        allValues = realExpenseTable[:, realExpenseTableHead.index('value')].astype(float)
        if None in allValues:
            allValues = allValues[allValues != np.array(None)]
        totalVal = sum(allValues)

        colTableName = realExpenseTable[:, realExpenseTableHead.index('table')]
        labels = []
        sizes = []
        for table in np.unique(colTableName):
            indx = np.where(realExpenseTable[:, realExpenseTableHead.index('table')]==table)
            smallArray = realExpenseTable[indx]
            values = sum(smallArray[:, realExpenseTableHead.index('value')].astype(float))
            txt = '{} = {:.2f}'.format(table, values)
            labels.append(txt)
            size = (values/totalVal)*100
            sizes.append(size)

        fig1, ax1 = plt.subplots()
        ax1.pie(sizes, labels=labels, autopct='%1.2f%%', startangle=90)
        ax1.axis('equal')
        plt.legend(title='Total: {:.2f}'.format(totalVal))

        plt.show()

    def plotNamePie(self):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        realExpenseTable, realExpenseTableHead = self.readPlanExpenses()
        allValues = realExpenseTable[:, realExpenseTableHead.index('value')].astype(float)
        if None in allValues:
            allValues = allValues[allValues != np.array(None)]
        totalVal = sum(allValues)

        colTableName = realExpenseTable[:, realExpenseTableHead.index('name')]
        labels = []
        sizes = []
        for table in np.unique(colTableName):
            indx = np.where(realExpenseTable[:, realExpenseTableHead.index('name')]==table)
            smallArray = realExpenseTable[indx]
            values = sum(smallArray[:, realExpenseTableHead.index('value')].astype(float))
            txt = '{} = {:.2f}'.format(table, values)
            labels.append(txt)
            size = (values/totalVal)*100
            sizes.append(size)

        fig1, ax1 = plt.subplots()
        ax1.pie(sizes, labels=labels, autopct='%1.2f%%', startangle=90)
        ax1.axis('equal')
        plt.legend(title='Total: {:.2f}'.format(totalVal))

        plt.show()

    def plotGraf(self):
        print('Module: {}, Class: {}, Def: {}'.format(__name__, __class__, sys._getframe().f_code.co_name))
        realExpenseTable, realExpenseTableHead = self.readPlanExpenses()
        planIncomeTable, planIncomeTableHead = self.readPlanIncome()
        x_exp = []
        y_exp = []
        for date in np.unique(realExpenseTable[:, realExpenseTableHead.index('payDay')]):
            indx = np.where(realExpenseTable[:, realExpenseTableHead.index('payDay')] == date)
            arr = realExpenseTable[indx, realExpenseTableHead.index('value')].astype(float)
            x_exp.append(date)
            y_exp.append(abs(sum(arr[0])))

        x_inc = []
        y_inc = []
        for date in np.unique(planIncomeTable[:, planIncomeTableHead.index('payDay')]):
            indx = np.where(planIncomeTable[:, planIncomeTableHead.index('payDay')] == date)
            arr = planIncomeTable[indx, planIncomeTableHead.index('value')].astype(float)
            x_inc.append(date)
            y_inc.append(abs(sum(arr[0])))

        fig1, ax1 = plt.subplots()
        ax1.plot(x_exp, y_exp)
        ax1.plot(x_inc, y_inc)
        # plt.setp(plt.get_xticklabels(), rotation=30, ha="right")
        fig1.autofmt_xdate()
        plt.grid()
        plt.show()


def main():
    app = QApplication(sys.argv)
    window = MyApp()
    window.show()
    # sys.exit(app.exec_())
    app.exec_()


if __name__ == '__main__':
    main()



