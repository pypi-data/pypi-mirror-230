#!/usr/bin/env python3
# -*- coding: utf-8 -*
"""Excel reader
"""

from pathlib import Path
import openpyxl
from .model import ExcelConfig
from .model import BookData

# 支持的数据类型 String, Number, StringArray, NumberArray


def read_excel_values(input_file, sheet_index=0):
    """read_excel_values
    Args:
        input_file (str): 输入 Excel 文件路径
        sheet_index (int): 工作簿序号. Defaults to 0.
        config (ExcelConfig): Excel 配置. Defaults to None.
    Raises:
        Exception: Excel 数据异常
    """
    input_file = Path(input_file)
    workbook = openpyxl.load_workbook(input_file, read_only=True, data_only=True, rich_text=False)
    sheet = workbook[workbook.sheetnames[sheet_index]]
    max_col = sheet.max_column
    max_row = sheet.max_row
    values = []
    i = 0
    for row in sheet.values:
        row_list = []
        for j, value in enumerate(row):
            if i == 0 and value is None:
                max_col = j
            if j == 0 and value is None:
                max_row = i
            if i >= max_row or j >= max_col:
                break
            row_list.append(_strip_str(value))
        if i >= max_row:
            break
        values.append(row_list)
        i += 1
    workbook.close()
    if max_row < 4:
        raise ValueError(f"ExcelError 数据最少不少于 4 行！{input_file.name}")
    return values


def read_excel_book(input_file, sheet_index=0, config: ExcelConfig = None) -> BookData:
    """read_excel_book
    Args:
        input_file (str): 输入 Excel 文件路径
        sheet_index (int): 工作簿序号. Defaults to 0.
        config (ExcelConfig): Excel 配置. Defaults to None.
    Raises:
        Exception: Excel 数据异常
    """
    _sheet_index = sheet_index
    _config = config or ExcelConfig()
    _input_file = Path(input_file)
    _data_name = _input_file.stem.strip()
    # 读取 Excel
    _values = read_excel_values(_input_file, sheet_index)
    # 属性名
    _keys = _values[_config.key_line - 1]
    del_cols = []
    for i, key in enumerate(_keys[::-1]):
        if not (len(key) > 0 and key[0].isalpha() and not key.startswith("ignore_")):
            del_cols.append(i)
    if len(del_cols) > 0:
        for row in _values:
            for i in del_cols:
                row.remove(i)
        # 删除指定列
        _values = [[row[i] for i in range(len(row)) if i not in del_cols] for row in _values]

    # 属性名
    _keys = [str(x) for x in _values[_config.key_line - 1]]
    # 标题描述
    _titles = [_strip_str(x) for x in _values[_config.title_line - 1]]
    # 类型定义
    _types = _values[_config.type_line - 1]
    # 值列表
    _values = _values[_config.value_line - 1 :]

    _keys_map = {}
    _repeat_index = 1
    for i, key in enumerate(_keys):
        if key in _keys_map:
            _repeat_index += 1
            _key2 = f"{key}{_repeat_index}"
            _keys[i] = _key2
            print("\033[1;31m" + f"[ValueError] {_data_name} 属性名重复！key: {key} 自动重命名为 {_key2}\033[0m")
        _keys_map[_keys[i]] = True
    _keys_map = None

    _schema = {_keys[i]: {"key": _keys[i], "type": _types[i], "title": _titles[i]} for i in range(len(_keys))}
    for i, row in enumerate(_values):
        for j, value in enumerate(row):
            line_info = f"{_data_name} => key: {_keys[j]}, line: {_config.value_line + i}"
            row[j] = _convert_value(value, _schema[_keys[j]]["type"], line_info)
    # BookData
    return BookData(_data_name, _schema, _keys, _values)


def _strip_str(value):
    if value is None:
        return ""
    text = str(value).strip()
    # 特殊空格和制表符替换为普通空格 \u00A0=不间断空格
    for space in ["\t", "\v", "\f", "\u3000", "\u00A0", "\u200F"]:
        text = text.replace(space, " ")
    # 移除不可见空格
    text = text.replace("\u200B", "")
    # 替换换行符
    text = text.replace("\r", "\n")
    text = text.replace("\n\n", "\n")
    text = text.replace("\n", "\\n")
    text = text.replace('"', '\\"')
    # 法语中叹号前面的空格替换为不间断空格, 防止标点换行到行首
    text = text.replace(" !", "\u00A0!")
    text = text.replace(" .", "\u00A0.")
    return text


def _convert_value(value, valueType, info=""):
    text = str(value).strip()
    if text.count('"') % 2 == 1:
        print("\033[1;31m" + f"[ValueError] 包含奇数个双引号! {info}, type: {valueType}, value: {value}" + "\033[0m")
    if valueType.lower() == "string":
        try:
            float_num = float(text)
            int_num = int(float_num)
            if float_num == int_num:
                text = str(int_num)
        except ValueError:
            pass
        return text
    elif valueType.lower() == "number":
        try:
            float_num = float(text)
            int_num = int(float_num)
            if float_num == int_num:
                return int_num
            return float_num
        except ValueError:
            if text.lower() == "true":
                return 1
            if text.lower() == "false":
                return 0
    elif valueType.lower().endswith("array"):
        ttype2 = valueType.replace("array", "").replace("Array", "")
        array = []
        for value2 in text.split(","):
            array.append(_convert_value(value2, ttype2, info))
        return array
    print("\033[1;31m" + f"[TypeError] 数据类型不匹配！{info}, type: {valueType}, value: {value}" + "\033[0m")
    return "error()"
